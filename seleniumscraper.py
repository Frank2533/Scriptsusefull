import json
import csv
import os
import zipfile
import openpyxl
import selenium
from selenium.webdriver.common.by import By
from selenium import webdriver
import configparser
import datetime
from time import sleep
import pickle
from skudataclass import skuunit
import selenium.common.exceptions
import time as _time
import pandas as pd
import concurrent.futures
import websites_scrape as ws
import traceback
import threading

Total_skus = []

date_time = str(datetime.datetime.now())
date, time = tuple(date_time.split(" "))
time = time.split(".")[0]
time = time.replace(":", "-")
page2ntf = 'SKU not scraped and to be done manually are '

file = open('Domaindatafile', 'rb')
domdatadict = pickle.load(file)



def get_config():
    global config
    config = configparser.ConfigParser()
    config.read(r'E:\Akash Hajnale\Google_scraper\config.ini')
    return config


def get_file(config):
    file_name = config['FILE']['file_name']
    row_num_start = int(config['FILE']['row_num_start'])
    column_num = int(config['FILE']['column_num'])
    return file_name, row_num_start, column_num


def get_proxy_var(config, country):
    Nohodofile = "E:\\Akash Hajnale\\Google_scraper\\Nohododatafile"
    file = open(Nohodofile, "rb")
    proxydict = pickle.load(file)

    COUNTRY = country
    if COUNTRY not in proxydict.keys():
        raise ValueError

    file.close()
    PROXY_HOST = config['PROXY']['PROXY_HOST']
    PROXY_PORT = proxydict[COUNTRY]
    PROXY_USER = config['PROXY']['PROXY_USER']
    PROXY_PASS = config['PROXY']['PROXY_PASS']
    # Proxy = f'https://{PROXY_USER}:{PROXY_PASS}@{PROXY_HOST}:{PROXY_PORT}'

    return COUNTRY, PROXY_HOST, PROXY_PORT, PROXY_USER, PROXY_PASS


def get_proxy_var_wonder(region, country):

    COUNTRY = country
    cstring = region.replace(" ", "").lower()
    PROXY_HOST = cstring + ".wonderproxy.com"
    PROXY_PORT = 11000
    PROXY_USER = "sachin"
    PROXY_PASS = "sachin@123"

    return COUNTRY, PROXY_HOST, PROXY_PORT, PROXY_USER, PROXY_PASS


def get_driver(PROXY_HOST=None, PROXY_PORT=None, PROXY_USER=None, PROXY_PASS=None, g_scrnsht_location=None, plugin=None):
    opt = webdriver.ChromeOptions()
    if plugin is None:
        manifest_json = """
           {
               "version": "1.0.0",
               "manifest_version": 2,
               "name": "Chrome Proxy",
               "permissions": [
                   "proxy",
                   "tabs",
                   "unlimitedStorage",
                   "storage",
                   "<all_urls>",
                   "webRequest",
                   "webRequestBlocking"
               ],
               "background": {
                   "scripts": ["background.js"]
               },
               "minimum_chrome_version":"22.0.0"
           }
           """

        background_js = """
           var config = {
                   mode: "fixed_servers",
                   rules: {
                     singleProxy: {
                       scheme: "http",
                       host: "%s",
                       port: parseInt(%s)
                     },
                     bypassList: ["localhost"]
                   }
                 };
    
           chrome.proxy.settings.set({value: config, scope: "regular"}, function() {});
    
           function callbackFn(details) {
               return {
                   authCredentials: {
                       username: "%s",
                       password: "%s"
                   }
               };
           }
    
           chrome.webRequest.onAuthRequired.addListener(
                       callbackFn,
                       {urls: ["<all_urls>"]},
                       ['blocking']
           );
           """ % (PROXY_HOST, PROXY_PORT, PROXY_USER, PROXY_PASS)
        pluginfile = 'proxy_auth_plugin.zip'
        with zipfile.ZipFile(pluginfile, 'w') as zp:
            zp.writestr("manifest.json", manifest_json)
            zp.writestr("background.js", background_js)
        opt.add_extension(pluginfile)

    else:
        pluginfile = plugin
        opt.add_extension(pluginfile)

    settings = {
        "recentDestinations": [{
            "id": "Save as PDF",
            "origin": "local",
            "account": ""
        }],
        "selectedDestinationId": "Save as PDF",
        "version": 2,
        "isHeaderFooterEnabled": True,
        "scalingTypePdf": 3,
        "isCssBackgroundEnabled": True
    }

    prefs = {
        'printing.print_preview_sticky_settings.appState': json.dumps(settings),
        'savefile.default_directory': g_scrnsht_location,
        "translate_whitelists": {"am": "en", "ar": "en", "eu": "en", "bn": "en", "en-GB": "en", "pt-BR": "en",
                                 "bg": "en",
                                 "ca": "en", "chr": "en", "hr": "en", "cs": "en", "da": "en", "nl": "en", "et": "en",
                                 "fil": "en", "fi": "en", "fr": "en", "de": "en", "el": "en", "gu": "en", "iw": "en",
                                 "hi": "en",
                                 "hu": "en", "is": "en", "id": "en", "it": "en", "ja": "en", "kn": "en", "ko": "en",
                                 "lv": "en",
                                 "lt": "en", "ms": "en", "ml": "en", "mr": "en", "no": "en", "pl": "en", "pt-PT": "en",
                                 "ro": "en",
                                 "ru": "en", "sr": "en", "zh-CN": "en", "sk": "en", "sl": "en", "es": "en", "sw": "en",
                                 "sv": "en",
                                 "ta": "en", "te": "en", "th": "en", "zh-TW": "en", "tr": "en", "ur": "en", "uk": "en",
                                 "vi": "en", "cy": "en"},
        "translate": {"enabled": "true"}
    }
    opt.add_experimental_option('prefs', prefs)
    opt.add_argument('--kiosk-printing')
    config = get_config()
    path = config.get('CHROME', 'path')
    driver = webdriver.Chrome(executable_path=path,
                              chrome_options=opt)


    return driver


def get_websites(g_scrnsht_location, skuunits, x, COUNTRY, PROXY_HOST, PROXY_PORT, PROXY_USER, PROXY_PASS):
    t = os.path.join(f"E:\\Akash Hajnale\\Google_scraper\\output_dir\\{COUNTRY}_{date}_{time}", "WEB_PAGES")
    os.mkdir(t)
    inactive_list = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
        for unit in skuunits:
            executor.submit(get_pages, g_scrnsht_location, x, unit, PROXY_HOST, PROXY_PORT, PROXY_USER, PROXY_PASS)
    for unit in skuunits:
        if unit.domainstat == "inactive":
            inactive_list.append(unit.pos)
            continue
        # print("Writing....")
    for one in range(0, len(inactive_list)):
        popitem = skuunits.pop(max(inactive_list) - 1)
        inactive_list.remove(max(inactive_list))
        print(f"found {popitem} inactive hence deleting")
    list_of_files_names = os.listdir(g_scrnsht_location)
    list_of_files_path = [os.path.join(g_scrnsht_location, x) for x in list_of_files_names]
    skupath = os.path.join(t, x.split("\"")[1])
    for g, v in zip(list_of_files_path, list_of_files_names):
        if ("Google" not in v.split(" ")) and ("Google.pdf" not in v.split(" ")):
            try:
                os.replace(g, os.path.join(skupath, v))
            except PermissionError:
                sleep(6)
                os.replace(g, os.path.join(skupath, v))


def get_proxy(config, proxy):
    pchoice, region, country = tuple(proxy)
    if pchoice is "N":
        COUNTRY, PROXY_HOST, PROXY_PORT, PROXY_USER, PROXY_PASS = get_proxy_var(config, country)
    else:
        COUNTRY, PROXY_HOST, PROXY_PORT, PROXY_USER, PROXY_PASS = get_proxy_var_wonder(region, country)
    # else:
    #     print("Wrong choice entered run program again")
    #     raise ValueError
    return COUNTRY, PROXY_HOST, PROXY_PORT, PROXY_USER, PROXY_PASS


def get_domdata(skuunit):
    global domdatadict
    if skuunit.domain in domdatadict.keys():
        if domdatadict[skuunit.domain]['TypeofWeb'] == "Retailer":
            skuunit.type = domdatadict[skuunit.domain]['TypeofWeb']
            skuunit.countryseller = domdatadict[skuunit.domain]['Country_seller']
            skuunit.sellercontact = domdatadict[skuunit.domain]['Seller_contact_Details']
            skuunit.sellercity = domdatadict[skuunit.domain]['City']
            skuunit.sellerstate = domdatadict[skuunit.domain]['State']
            skuunit.sellername = domdatadict[skuunit.domain]['SELLER_NAME']

        else:
            skuunit.type = domdatadict[skuunit.domain]['TypeofWeb']
            skuunit.countryseller = None
    else:
        pass


def get_skus(file_name, row_num_start, column_num):
    SKUlist = []
    wb_obj = openpyxl.load_workbook(file_name)
    sheet_obj = wb_obj.active
    row_num = row_num_start  # from config_file
    column_num = column_num  # from config_file
    cell_obj = sheet_obj.cell(row=row_num, column=column_num)
    print("The SKU's found are:\n")
    while cell_obj.value != None:
        print(cell_obj.value)
        SKUlist.append("\"" + str(cell_obj.value) + "\"")
        row_num += 1
        cell_obj = sheet_obj.cell(row=row_num, column=column_num)

    print(f"""

    ####################### TOTAL SKU's FOUND = {row_num - row_num_start} ########################

    """)
    return SKUlist


def get_pages(g_scrnsht_location, SKU, unit, PROXY_HOST, PROXY_PORT, PROXY_USER, PROXY_PASS):
    driver = get_driver(PROXY_HOST, PROXY_PORT, PROXY_USER, PROXY_PASS, g_scrnsht_location)
    link = unit.url
    pos = unit.pos
    status = "active"
    comment = None
    remark = None
    print(f"Website {pos} loading.. ")
    SKU = SKU.split("\"")[1]
    print(SKU)
    w = os.path.join(os.getcwd(), "WEB_PAGES")
    if pos == 1:
        w = os.path.join(w, SKU)
        os.mkdir(w)
    try:
        driver.get(link)
        title_check = ["404", "403", "402", "Access Denied"]

        if [ele for ele in title_check if (ele in driver.title)]:
            print("page not found")
            status = "inactive"

    except:
        try:
            # driver.delete_all_cookies()
            driver.refresh()
        except:
            status = "inactive"
            print("Connection problem")

    # if status is "active":
        # try:
        #     # status, comment = get_sku_status(driver.page_source, SKU, status, comment)
        # except selenium.common.exceptions.TimeoutException:
        #     print("Check this site manually")
        #     remark = "Check this site manually"

    if status == "active":
        try:
            driver.execute_script("window.print();")
        except:
            print(f"{link} not printed")
            remark = "not printed"
    driver.quit()
    unit.domainstat = status
    unit.domaincomment = comment
    unit.remark = remark


def select_country(driver, country):
    while True:
        try:
            driver.get("https://www.google.com/preferences?hl=en&fg=1")
            driver.find_element(By.CSS_SELECTOR, "a#regionanchormore.jfk-link").click()
            country_list = driver.find_elements(By.CSS_SELECTOR, "div.DB6WRb")
            break
        except:
            driver.refresh()

        # print("countrylist", country_list)
    for countr in country_list:
        cn = countr.find_element(By.CSS_SELECTOR, "span.jfk-radiobutton-label")
        # print(countr)
        if cn.text == country:
            vi = 1
            cn.click()
            print(f"Selected {cn.text} as country")
            break
        # else:
        #     print(f"***************!!!!!COUNTRY {country} NOT FOUND!!!!!******************")
    if vi != 1:
        print(f"Country {country} not found")
    driver.find_element(By.CSS_SELECTOR, "div.goog-inline-block.jfk-button.jfk-button-action").click()



def scrape_data(project_name,proxy_list, proxyno, country = None):

    config = get_config()
    global Total_skus
    global projectpath
    global week
    global page2ntf

    COUNTRY, PROXY_HOST, PROXY_PORT, PROXY_USER, PROXY_PASS = get_proxy(config, proxy_list[proxyno])
    if country != None:
        COUNTRY = country
    threading.current_thread().setName(f"{COUNTRY}")
    date_time = str(datetime.datetime.now())
    date, time = tuple(date_time.split(" "))
    time = time.split(".")[0]
    time = time.replace(":", "-")
    report_path = os.path.join(projectpath, f"{COUNTRY}_{date}_{time}_{proxyno + 1}")
    os.mkdir(report_path)
    g_scrnsht_location = os.path.join(f"{report_path}",
                     "Google_scrnshts")
    os.mkdir(g_scrnsht_location)
    os.chdir(report_path)

    file_name, row_num_start, column_num = get_file(config)
    skulist = get_skus(file_name, row_num_start, column_num)

    filename = os.path.join(report_path, "Final Reports")
    os.mkdir(filename)
    fields = ["SKU","SKU_Type", "Country_Seller", "Domain", "URL",
              "Country_monitored", "Pos", "Review_date",
              "type", "Sellername","VAT/GST", "Partnerproid",
              "Sellercontact", "SellerCity", "SellerState",
              "Domain_Status", "Domain_comment",
              "Week", "Remark_during_program"]
    filename = f"{filename}\\{COUNTRY}_{date}_{time}_FINAL.csv"
    csvfile = open(filename, "a", encoding="utf-8")
    csvwriter = csv.writer(csvfile)
    csvwriter.writerow(fields)
    filename = f"{projectpath}\\Total_skuslistdata_{proxyno + 1}"
    file = open(filename, "wb")
    skusforcountry = []


        # driver.switch_to.alert.accept()


    for count, x in enumerate(skulist, start=1):
        while True:
            try:
                driver = get_driver(PROXY_HOST, PROXY_PORT, PROXY_USER, PROXY_PASS, g_scrnsht_location)
                break
            except Exception as e:
                print(e)


        if country is not None:
            select_country(driver, country)


        driver.set_page_load_timeout(60)

        skuunits = []
        ck = 0
        pos = 1
        try:
            driver.get("http://www.google.com")
        except selenium.common.exceptions.WebDriverException:
            sleep(3)
            driver.refresh()

        while True:
            try:
                element = driver.find_element_by_name("q")
                break
            except selenium.common.exceptions.NoSuchElementException:
                sleep(3)
                driver.refresh()
                try:
                    element = driver.find_element_by_name("q")
                except selenium.common.exceptions.NoSuchElementException:
                    print("no such element error occured")
                    driver.refresh()

        while True:
            try:
                element.send_keys(x)
                break
            except:
                driver.find_element(By.CSS_SELECTOR, "button#L2AGLb").click()

        element.submit()
        sleep(3)

        print(f"Gathering SKU({x}) count: {count} ")
        if count % 5 == 0:
            sleep(2)
        results = driver.find_elements(By.CSS_SELECTOR, "div.yuRUbf a")
        if count % 3 == 0:
            sleep(1)

        driver.execute_script("window.print();")

        for result in results:
            url = result.get_attribute("href")
            dom = url.split("/")[2]
            if dom == "translate.google.com":
                continue
            if dom == "webcache.googleusercontent.com":
                continue
            if dom.split(".")[0] == "www":
                dom = dom.split(".")
                del dom[0]
                dom = (".").join(dom)
            skuunits.append(skuunit(x.split("\"")[1], pos, COUNTRY, url, dom, week))
            pos += 1

        try:
            driver.find_element(By.CSS_SELECTOR, "a#pnnext").click()

        except:
            sleep(36)
            print(f"{x} page 2 not scraped")
            page2ntf = page2ntf + f':{COUNTRY} - {x}'
            continue

        try:
            results = driver.find_elements(By.CSS_SELECTOR, "div.yuRUbf a")
        except:
            sleep(5)
            driver.refresh()

        for result in results:
            url = result.get_attribute("href")
            dom = url.split("/")[2]
            if dom == "translate.google.com":
                continue
            if dom == "webcache.googleusercontent.com":
                continue
            if dom.split(".")[0] == "www":
                dom = dom.split(".")
                del dom[0]
                dom = (".").join(dom)
            skuunits.append(skuunit(x.split("\"")[1], pos, COUNTRY, url, dom, week))
            pos += 1
        #
        # if count % 4 == 0:
        #     try:
        #         (driver.find_elements(By.CSS_SELECTOR, "h3.LC20lb.MBeuO.DKV0Md"))[2].click()
        #     except:
        #         break
        #
        # elif count % 3 == 0:
        #     try:
        #         (driver.find_elements(By.CSS_SELECTOR, "h3.LC20lb.MBeuO.DKV0Md"))[1].click()
        #     except:
        #         break
        #
        # elif count % 8 == 0:
        #     try:
        #         (driver.find_elements(By.CSS_SELECTOR, "h3.LC20lb.MBeuO.DKV0Md"))[4].click()
        #     except:
        #         break
        print(f"Total {len(skuunits)} webpages found")
        print(skuunits)

        # get_websites(u, skuunits, x, COUNTRY, PROXY_HOST, PROXY_PORT, PROXY_USER, PROXY_PASS)

        for unit in skuunits:

            unit.proxyused = proxy_list[proxyno]
            get_domdata(unit)
            if unit.type == "Marketplace":
                try:
                    ws.get_country(unit)
                    get_marketplace_data(unit, driver)
                except Exception as e:
                    print(f"\n\nException occured while scraping {unit.domain}\n\n")
                    print(traceback.format_exc())

            # "Sellername", "VAT/GST", "Partnerproid", "Sellercontact", "SellerCity", "SellerState"
            fields = [unit.sku, unit.SKUtype, unit.countryseller,
                      unit.domain, unit.url, unit.country, unit.pos,
                      unit.reviewdate, unit.type, unit.sellername,
                      unit.vatgstnumber, unit.parterproid, unit.sellercontact,
                      unit.sellercity, unit.sellerstate,
                      unit.domainstat, unit.domaincomment, unit.week, unit.remark]
            csvwriter.writerow(fields)
            Total_skus.append(unit)
            skusforcountry.append(unit)
        driver.quit()
    pickle.dump(skusforcountry, file, protocol=pickle.HIGHEST_PROTOCOL)
    file.close()
    csvfile.close()
    return f"{COUNTRY} completed"

        # if count % 15 == 0:
        #     print("Sleep in progress - 3min")
        #     sleep(180)


def scrape():

    global date
    global time
    global week
    proxy_list = []
    project_name = str(input("Enter the project name: "))
    week = int(input("Enter week no. : "))
    noofreports = int(input("Enter the number of reports required: "))
    # project_name, week, noofreports = ("test", 1, 1)
    project_name = project_name + f"_{date}_{time}"
    week = f"Week {week} {date.split('-')[0]}"
    pch = input("You wish to use proxy?(Y/N)")
    print("""Enter the proxy and country in following format:
                                   'N'- for nohodo proxy, 'W'- for wonderproxy
                                    Refer the excel files for region or country names
                                    FORMAT = 'proxy':'region':'country'
                                    EXAMPLE = W:Sydney:Australia
                                              N:None:Taiwan
                                    Check the proxy excels for more info.."""
          )



    project_name = project_name + f"_{noofreports}"
    global projectpath
    global cwd
    projectpath = os.path.join(cwd, f"output_dir")
    projectpath = os.path.join(projectpath, f"{project_name}")
    os.mkdir(projectpath)
    logpath = os.path.join(projectpath, "log.txt")
    # sys.stdout = open(logpath, 'w')

    if pch == 'Y':
        for x in range(1, noofreports+1):
            country = input(f"Proxy {x} : ")
            # country = "W:sydney:Australia"
            proxy_list.append(country.split(":"))
        print("Working on it....\n                      Sit back & relax....")
        # if noofreports <= 4:
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:

            result = [executor.submit(scrape_data, project_name, proxy_list, proxyno) for proxyno in range(0, noofreports) if sleep(5) is None]
        for r in result:
            try:
                print(r.result())
            except Exception as e:
                print(traceback.format_exc())
                print("Exception occured!!!" + str(e))

    else:
        country_list = []
        proxy = input('Enter proxy to be used: ').split(":")
        proxy_list.append(proxy)
        for x in range(1, noofreports+1):
            country = input(f"Enter country {x}: ")
            country_list.append(country)
        print("Working on it....\n                      Sit back & relax....")
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:

            result = [executor.submit(scrape_data, project_name, proxy_list,0, country_list[proxyno]) for proxyno in
                      range(0, noofreports) if sleep(5) is None]
        for r in result:
            try:
                print(r.result())
            except Exception as e:
                print(traceback.format_exc())
                print("Exception occured!!!" + str(e))

    filename = f"{projectpath}\\Total_skus_unfiltered.csv"

    csvfile = open(filename, "a", encoding="utf-8")
    fields = ["SKU","SKU_Type", "Country_Seller", "Domain", "URL",
              "Country_monitored", "Pos", "Review_date",
              "type", "Sellername","VAT/GST", "Partnerproid",
              "Sellercontact", "SellerCity", "SellerState",
              "Domain_Status", "Domain_comment",
              "Week", "Remark_during_program", "Proxy_used"]
    csvwriter = csv.writer(csvfile)
    csvwriter.writerow(fields)
    for unit in Total_skus:
        fields = [unit.sku, unit.SKUtype, unit.countryseller,
                      unit.domain, unit.url, unit.country, unit.pos,
                      unit.reviewdate, unit.type, unit.sellername,
                      unit.vatgstnumber, unit.parterproid, unit.sellercontact,
                      unit.sellercity, unit.sellerstate,
                      unit.domainstat, unit.domaincomment,
                      unit.week, unit.remark, unit.proxyused]
        csvwriter.writerow(fields)
    filename = f"{projectpath}\\Total_skuslistdata"
    file = open(filename, "wb")
    pickle.dump(Total_skus, file, protocol=pickle.HIGHEST_PROTOCOL)
    file.close()
    return projectpath


def create_common_sku(projectpath):
    dtype = {"SKU":str,"SKU_Type":str, "Country_Seller":str, "Domain":str, "URL":str,
              "Country_monitored":str, "Pos":int, "Review_date":str,
              "type":str, "Sellername":str,"VAT/GST":str, "Partnerproid":str,
              "Sellercontact":str, "SellerCity":str, "SellerState":str,
              "Domain_Status":str, "Domain_comment":str,
              "Week":str, "Remark_during_program":str, "Proxy_used":str}
    table = pd.read_csv(projectpath + "\\Total_skus_unfiltered.csv",
                        dtype=dtype)
    common_table = table.groupby(['SKU','type', 'URL', 'Domain'])['Pos', 'Country_monitored', 'Proxy_used'].agg(';'.join).reset_index()
    common_table.to_csv(projectpath+"\\Common+unique_skus.csv", index=False)


def dump_projectpath(projectpath):
    config = get_config()
    config.set('PROJECTPATH','path',projectpath)
    with open('config.ini', 'w') as configfile:
        config.write(configfile)

def get_marketplace_data(skuunit, driver):
    dom_dict = {"akakce.com":"", "alibaba.com":"get_alibaba_com", "alliedcomputersng.com":"", "amazon.ca":"", "amazon.co.jp":"",
             "amazon.co.uk":"", "amazon.com":"", "amazon.com.au":"", "amazon.de":"", "amazon.eg":"", "amazon.sa":"",
             "bazaar.shopclues.com":"", "blibli.com":"", "bol.com":"", "bukalapak.com":"", "cdw.ca":"",
             "chinese.alibaba.com":"", "ebay.co.uk":"", "ebay.com":"get_ebay_com", "ebay.com.au":"get_ebay_com_au", "ebay.de":"",
             "ebay.it":"", "global.microless.com":"", "govets.com":"", "id.aliexpress.com":"", "indiamart.com":"get_indiamart_com",
             "indonesian.alibaba.com":"", "ja.aliexpress.com":"", "jd.co.th":"", "jd.id":"", "judge.me":"", "konga.com":"",
             "korean.alibaba.com":"", "lazada.co.id":"", "lazada.co.th":"", "lazada.com":"", "lazada.com.my":"get_lazada_com_my",
             "lazada.com.ph":"", "lazada.sg":"", "mall.shopee.co.id":"", "padiumkm.id":"", "picclick.com":"",
             "pl.aliexpress.com":"", "progressivepii.com":"", "retif.eu":"", "ruten.com.tw":"", "shopclues.com":"",
             "shopee.co.id":"", "shopee.co.th":"", "shopee.com":"", "shopee.com.br":"", "shopee.com.my":"",
             "shopee.ph":"", "shopee.pn":"", "shopee.sg":"", "shopee.tw":"", "shopee.vn":"", "spanish.alibaba.com":"",
             "stock.kachaf.com":"", "sugarworkshop.en.alibaba.com":"", "thai.alibaba.com":"", "tokopedia.com":"get_tokopedia_com",
            "tradeindia.com":"", "ubuy.co.in":"", "walmart.com":"", "www.pchome.co.th":"", "aliexpress.com":"get_aliexpress_com" }
    if skuunit.domain in dom_dict.keys():
        if dom_dict[skuunit.domain]!="":
            fun_name = dom_dict[skuunit.domain]
            eval(f"ws.{fun_name}(skuunit, driver)")
        else:
            pass
    else:
        pass


if __name__ == "__main__":

    global cwd
    cwd = os.getcwd()
    start_time = _time.time()
    projectpath = scrape()
    print(page2ntf)
    # projectpath = "E:\Akash Hajnale\Google_scraper\output_dir\Week_10_2022-07-01_16-43-13_2022-07-01_16-43-13_7"
    create_common_sku(projectpath)
    dump_projectpath(projectpath)
    print("--- %s seconds ---" % round(_time.time() - start_time))


